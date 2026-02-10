"""
Excel export module for PDF analysis results.

Creates formatted Excel files with auto-sized columns and styled headers.
"""

from typing import List, Dict, Any
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def create_excel_file(
    results: List[Dict[str, Any]],
    column_names: List[str]
) -> bytes:
    """
    Create an Excel file from analysis results.

    Args:
        results: List of result dictionaries, each containing analysis data
        column_names: Ordered list of column names (including "Document Name")

    Returns:
        Excel file content as bytes
    """
    # Create workbook and active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "PDF Analysis Results"

    # Write headers
    _write_headers(ws, column_names)

    # Write data rows
    _write_data_rows(ws, results, column_names)

    # Auto-size columns
    _auto_size_columns(ws, column_names, results)

    # Save to bytes
    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)

    return excel_bytes.getvalue()


def _write_headers(ws, column_names: List[str]) -> None:
    """
    Write and style header row.

    Args:
        ws: openpyxl worksheet
        column_names: List of column names
    """
    # Header styling
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_alignment = Alignment(horizontal="left", vertical="center")

    for col_idx, col_name in enumerate(column_names, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment


def _write_data_rows(
    ws,
    results: List[Dict[str, Any]],
    column_names: List[str]
) -> None:
    """
    Write data rows to worksheet.

    Args:
        ws: openpyxl worksheet
        results: List of result dictionaries
        column_names: Ordered list of column names
    """
    for row_idx, result_dict in enumerate(results, start=2):
        for col_idx, col_name in enumerate(column_names, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = result_dict.get(col_name, "")
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _auto_size_columns(
    ws,
    column_names: List[str],
    results: List[Dict[str, Any]]
) -> None:
    """
    Auto-size columns based on content width.

    Args:
        ws: openpyxl worksheet
        column_names: List of column names
        results: List of result dictionaries
    """
    for col_idx, col_name in enumerate(column_names, start=1):
        # Calculate max width needed
        max_width = len(col_name) + 2  # Start with header width

        # Check data rows
        for result_dict in results:
            value = str(result_dict.get(col_name, ""))
            # Handle multi-line content
            lines = value.split('\n')
            max_line_length = max(len(line) for line in lines) if lines else 0
            max_width = max(max_width, max_line_length + 2)

        # Cap maximum width at 50 for readability
        max_width = min(max_width, 50)

        # Set column width
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = max_width


def format_results_for_export(
    results: List[Dict[str, Any]],
    filenames: List[str],
    column_names: List[str]
) -> List[Dict[str, Any]]:
    """
    Format analysis results for Excel export by adding document names
    and ensuring all columns are present.

    Args:
        results: List of analysis result dictionaries
        filenames: List of corresponding PDF filenames
        column_names: All column names including "Document Name"

    Returns:
        Formatted list of dictionaries ready for Excel export
    """
    formatted_results = []

    for filename, result_dict in zip(filenames, results):
        formatted_row = {"Document Name": filename}

        # Add all other columns
        for col_name in column_names:
            if col_name != "Document Name":
                formatted_row[col_name] = result_dict.get(col_name, "")

        formatted_results.append(formatted_row)

    return formatted_results
